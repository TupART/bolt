import pandas as pd
import openpyxl
from app.utils.data_mapper import get_pcc_step2_data, get_step3_data
import os

def process_excel_data(file_path):
    df = pd.read_excel(file_path)
    data = []
    
    for idx, row in df.iterrows():
        name = row['Name']
        surname = row['Surname']
        market = row['Market']
        is_pcc = row['Va a ser PCC?'] == 'Y'
        is_ds = row['Va a ser PCC?'] == 'DS'
        is_tl = row['Va a ser PCC?'] == 'TL'
        windows_username = row['B2E User Name']
        
        step2_data = get_pcc_step2_data(windows_username, market, is_pcc, is_ds)
        step3_data = get_step3_data(market, is_pcc, is_ds, is_tl)
        
        data.append({
            'id': idx,
            'name': name,
            'surname': surname,
            'step2': step2_data,
            'step3': step3_data,
            'step4': {
                'name': name,
                'surname': surname,
                'email': row['E-mail'],
                'market': market,
                'pcc_status': row['Va a ser PCC?'],
                'b2e_username': windows_username
            }
        })
    
    return data

def process_step4_template(selected_rows):
    uploads_dir = 'uploads'
    file_path = os.path.join(uploads_dir, os.listdir(uploads_dir)[0])
    df = pd.read_excel(file_path)
    
    plantilla = 'PlantillaSTEP4.xlsx'
    wb = openpyxl.load_workbook(plantilla)
    ws = wb.active
    
    for row in selected_rows:
        idx = int(row)
        name = df.iloc[idx]['Name']
        surname = df.iloc[idx]['Surname']
        email = df.iloc[idx]['E-mail']
        market = df.iloc[idx]['Market']
        pcc_status = df.iloc[idx]['Va a ser PCC?']
        b2e_username = df.iloc[idx]['B2E User Name']
        
        row_num = 7 + idx
        ws[f'C{row_num}'] = name
        ws[f'D{row_num}'] = surname
        ws[f'E{row_num}'] = email
        
        if pcc_status == 'Y':
            if market == 'DACH':
                ws[f'F{row_num}'] = "/+4940210918145 /+43122709858 /+41445295828"
                ws[f'G{row_num}'] = "D_PCC"
                ws[f'H{row_num}'] = "Team_D_CCH_PCC_1"
            elif market == 'France':
                ws[f'F{row_num}'] = "/+33180037979"
                ws[f'G{row_num}'] = "F_PCC"
                ws[f'H{row_num}'] = "Team_F_CCH_PCC_1"
            elif market == 'Spain':
                ws[f'F{row_num}'] = "/+34932952130"
                ws[f'G{row_num}'] = "E_PCC"
                ws[f'H{row_num}'] = "Team_E_CCH_PCC_1"
            elif market == 'Italy':
                ws[f'F{row_num}'] = "/+390109997099"
                ws[f'G{row_num}'] = "I_PCC"
                ws[f'H{row_num}'] = "Team_I_CCH_PCC_1"
        
        ws[f'L{row_num}'] = "Y" if pcc_status == 'Y' else "N"
        ws[f'Q{row_num}'] = b2e_username
        ws[f'R{row_num}'] = b2e_username
        ws[f'V{row_num}'] = "Team Leader" if pcc_status == 'TL' else "Agent"
    
    output_file = 'PlantillaSTEP4_Rellenada.xlsx'
    wb.save(output_file)
    return output_file